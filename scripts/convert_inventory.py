import json
from pathlib import Path

import openpyxl

SOURCE = Path("/tmp/wine-cellar-export/Wine-Cellar-Twin/attached_assets/wine_inventory_with_correct_photo_filenames_1775163052338.xlsx")
OUTPUT = Path("/Users/athenalynn/Desktop/Ventures/ordering app/archive/country_club_pool_ordering_mvp/wine-cellar-rebuild/src/data/wines.json")

wb = openpyxl.load_workbook(SOURCE, data_only=True)
ws = wb["Inventory"]
headers = [cell.value for cell in ws[1]]

rows = []
counters = {
    (1, "top"): 0,
    (1, "bottom"): 0,
    (2, "top"): 0,
    (2, "bottom"): 0,
}
chilled = {"White", "Sparkling", "Dessert", "Rose", "Rosé"}

for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    record = dict(zip(headers, row))
    price = float(record.get("Estimated Bottle Value (USD)") or 0)
    category = str(record.get("Category") or "").strip()
    cellar = 1 if price < 50 else 2
    zone = "top" if category in chilled else "bottom"
    counters[(cellar, zone)] += 1

    vintage = record.get("Vintage")
    if vintage in (None, "NV"):
        vintage = None
    else:
        vintage = int(vintage)

    rows.append(
        {
            "id": idx,
            "producer": record.get("Producer"),
            "wineName": record.get("Wine"),
            "vintage": vintage,
            "region": record.get("Appellation / Region"),
            "country": record.get("Country"),
            "variety": record.get("Type / Variety"),
            "category": category,
            "size": record.get("Size"),
            "quantity": int(record.get("Qty") or 1),
            "estimatedPrice": price,
            "sourceUrl": record.get("Value Source URL"),
            "notes": record.get("Confidence / Notes"),
            "frontPhoto": record.get("Primary Photo File Name"),
            "backPhoto": record.get("Supporting Photo File Name(s)"),
            "cellar": cellar,
            "zone": zone,
            "slot": counters[(cellar, zone)],
            "status": "Ready" if vintage and vintage <= 2024 else "Hold",
            "tastingNotes": None,
            "drinkWindow": None,
        }
    )

OUTPUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(rows)} wines to {OUTPUT}")
print(counters)
